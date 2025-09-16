"""
Warehouse Transfer Planning Tool - Import/Export Module
Handles Excel and CSV data import/export with validation and formatting
"""

import pandas as pd
import numpy as np
from datetime import datetime
import io
import os
import logging
from typing import Dict, List, Tuple, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import pymysql
from pathlib import Path

# Import our modules
try:
    from . import database
    from . import models
    from . import calculations
except ImportError:
    # For direct execution
    import database
    import models
    import calculations

logger = logging.getLogger(__name__)

class ImportExportManager:
    """
    Comprehensive import/export manager for warehouse transfer data
    
    Handles:
    - Excel file import with validation
    - Excel export with professional formatting
    - CSV import/export capabilities
    - Data validation and error reporting
    - Transfer order generation
    """
    
    def __init__(self):
        """Initialize the import/export manager"""
        self.validation_errors = []
        self.validation_warnings = []
        
    # =============================================================================
    # EXCEL IMPORT FUNCTIONALITY
    # =============================================================================
    
    def import_excel_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Import Excel file with comprehensive validation and error handling
        
        Args:
            file_content: Raw bytes of the uploaded Excel file
            filename: Original filename for error reporting
            
        Returns:
            Dict containing import results, errors, and statistics
        """
        try:
            # Reset validation tracking
            self.validation_errors = []
            self.validation_warnings = []
            
            # Read Excel file into pandas DataFrame
            excel_data = pd.read_excel(io.BytesIO(file_content), sheet_name=None)
            
            results = {
                "success": False,
                "filename": filename,
                "import_timestamp": datetime.now().isoformat(),
                "sheets_processed": 0,
                "records_imported": 0,
                "errors": [],
                "warnings": [],
                "summary": {}
            }
            
            # Process each sheet
            for sheet_name, df in excel_data.items():
                sheet_result = self._process_excel_sheet(sheet_name, df)
                results["sheets_processed"] += 1
                results["records_imported"] += sheet_result.get("records", 0)
                
            # Compile validation results
            results["errors"] = self.validation_errors
            results["warnings"] = self.validation_warnings
            results["success"] = len(self.validation_errors) == 0
            
            # Generate import summary
            results["summary"] = self._generate_import_summary(results)
            
            logger.info(f"Excel import completed: {filename}, Success: {results['success']}")
            return results
            
        except Exception as e:
            logger.error(f"Excel import failed: {str(e)}")
            return {
                "success": False,
                "filename": filename,
                "error": f"Import failed: {str(e)}",
                "import_timestamp": datetime.now().isoformat()
            }
    
    def _process_excel_sheet(self, sheet_name: str, df: pd.DataFrame) -> Dict[str, Any]:
        """Process individual Excel sheet based on its content type"""
        
        sheet_name_lower = sheet_name.lower()
        
        if 'inventory' in sheet_name_lower or 'stock' in sheet_name_lower:
            return self._import_inventory_data(df, sheet_name)
        elif 'sales' in sheet_name_lower or 'history' in sheet_name_lower:
            return self._import_sales_data(df, sheet_name)
        elif 'sku' in sheet_name_lower or 'product' in sheet_name_lower:
            return self._import_sku_data(df, sheet_name)
        elif 'stockout' in sheet_name_lower:
            return self._import_stockout_data(df, sheet_name)
        else:
            # Try to auto-detect content type
            return self._auto_detect_and_import(df, sheet_name)
    
    def _import_inventory_data(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """
        Import inventory levels data from Excel sheet with comprehensive error logging

        Tracks each row's success/failure with specific error reasons and line numbers.
        Pre-validates SKUs against master table to prevent foreign key violations.
        Supports partial imports with detailed reporting.

        Args:
            df: DataFrame containing inventory data
            sheet_name: Name of the Excel sheet for error reporting

        Returns:
            Dict containing detailed import results with line-by-line tracking
        """

        # Required column is just sku_id, quantity columns are flexible
        required_columns = ['sku_id']
        result = {
            "type": "inventory",
            "records": 0,
            "errors": [],
            "import_summary": {
                "total_rows": 0,
                "successful": 0,
                "failed": 0,
                "skipped": 0
            },
            "import_details": []
        }

        # Check for sku_id column (required)
        missing_required = self._validate_required_columns(df, required_columns, sheet_name)
        if missing_required:
            return result

        # Check which quantity columns are available
        has_burnaby = any('burnaby' in col.lower() and 'qty' in col.lower() for col in df.columns)
        has_kentucky = any('kentucky' in col.lower() and 'qty' in col.lower() for col in df.columns)

        # At least one quantity column is required
        if not has_burnaby and not has_kentucky:
            self.validation_errors.append(
                f"Missing quantity columns in {sheet_name}: Need at least one of 'burnaby_qty' or 'kentucky_qty'"
            )
            return result

        # Create clean dataframe with available columns
        # Find the actual sku_id column name (case-insensitive)
        sku_id_col = next(col for col in df.columns if col.lower() == 'sku_id')
        columns_to_use = [sku_id_col]
        if has_burnaby:
            burnaby_col = next(col for col in df.columns if 'burnaby' in col.lower() and 'qty' in col.lower())
            columns_to_use.append(burnaby_col)
        if has_kentucky:
            kentucky_col = next(col for col in df.columns if 'kentucky' in col.lower() and 'qty' in col.lower())
            columns_to_use.append(kentucky_col)

        df_clean = df[columns_to_use].copy()
        df_clean = df_clean.dropna(subset=[sku_id_col])

        # Standardize column names and add missing columns with default values
        df_clean = df_clean.rename(columns={
            col: 'burnaby_qty' if 'burnaby' in col.lower() else 'kentucky_qty' if 'kentucky' in col.lower() else 'sku_id' if col.lower() == 'sku_id' else col
            for col in df_clean.columns
        })

        # Add missing quantity columns with default value 0
        if 'burnaby_qty' not in df_clean.columns:
            df_clean['burnaby_qty'] = 0
            self.validation_warnings.append(f"Burnaby quantities defaulted to 0 in {sheet_name}")
        if 'kentucky_qty' not in df_clean.columns:
            df_clean['kentucky_qty'] = 0
            self.validation_warnings.append(f"Kentucky quantities defaulted to 0 in {sheet_name}")

        # Convert quantities to integers and track conversion errors
        for qty_col in ['burnaby_qty', 'kentucky_qty']:
            df_clean[qty_col] = pd.to_numeric(df_clean[qty_col], errors='coerce').fillna(0).astype(int)

        # Track original row numbers (adding 2 to account for header row and 0-based indexing)
        df_clean['original_row'] = df_clean.index + 2

        result["import_summary"]["total_rows"] = len(df_clean)

        # Pre-load all existing SKUs from master table and inventory table
        try:
            db = database.get_database_connection()
            cursor = db.cursor()

            # Get all SKUs that exist in master table
            cursor.execute("SELECT sku_id FROM skus")
            existing_master_skus = {row[0] for row in cursor.fetchall()}

            # Get all SKUs that exist in inventory table
            cursor.execute("SELECT sku_id FROM inventory_current")
            existing_inventory_skus = {row[0] for row in cursor.fetchall()}

            logger.info(f"Found {len(existing_master_skus)} SKUs in master table, {len(existing_inventory_skus)} in inventory table")

        except Exception as e:
            self.validation_errors.append(f"Failed to load existing SKUs: {str(e)}")
            logger.error(f"Database error loading SKUs: {str(e)}")
            db.close()
            return result

        # Track duplicates within the CSV file
        sku_counts = df_clean['sku_id'].value_counts()
        duplicate_skus = set(sku_counts[sku_counts > 1].index)

        # Process each row with detailed tracking
        imported_count = 0

        for idx, row in df_clean.iterrows():
            line_number = row['original_row']
            sku_id = row['sku_id']

            detail_entry = {
                "line_number": line_number,
                "sku_id": sku_id,
                "burnaby_qty": row['burnaby_qty'],
                "kentucky_qty": row['kentucky_qty'],
                "status": "pending",
                "error_category": None,
                "error_message": None
            }

            try:
                # Check for duplicate SKU in file
                if sku_id in duplicate_skus:
                    detail_entry["status"] = "failed"
                    detail_entry["error_category"] = "DUPLICATE_IN_FILE"
                    detail_entry["error_message"] = f"SKU {sku_id} appears multiple times in CSV file"
                    result["import_summary"]["failed"] += 1
                    result["import_details"].append(detail_entry)
                    continue

                # Validate SKU exists in master table
                if sku_id not in existing_master_skus:
                    detail_entry["status"] = "failed"
                    detail_entry["error_category"] = "SKU_NOT_IN_MASTER"
                    detail_entry["error_message"] = f"SKU {sku_id} does not exist in master SKU table"
                    result["import_summary"]["failed"] += 1
                    result["import_details"].append(detail_entry)
                    continue

                # Validate data quality
                if row['burnaby_qty'] < 0 or row['kentucky_qty'] < 0:
                    detail_entry["status"] = "failed"
                    detail_entry["error_category"] = "INVALID_DATA"
                    detail_entry["error_message"] = f"Negative quantities not allowed (Burnaby: {row['burnaby_qty']}, Kentucky: {row['kentucky_qty']})"
                    result["import_summary"]["failed"] += 1
                    result["import_details"].append(detail_entry)
                    continue

                # Check which columns were actually provided
                update_burnaby = has_burnaby
                update_kentucky = has_kentucky
                update_both = update_burnaby and update_kentucky

                # Check if SKU exists in inventory table
                if sku_id in existing_inventory_skus:
                    # Update existing inventory record
                    if update_both:
                        query = """
                        UPDATE inventory_current
                        SET burnaby_qty = %s, kentucky_qty = %s, last_updated = NOW()
                        WHERE sku_id = %s
                        """
                        cursor.execute(query, (row['burnaby_qty'], row['kentucky_qty'], sku_id))
                    elif update_burnaby:
                        query = """
                        UPDATE inventory_current
                        SET burnaby_qty = %s, last_updated = NOW()
                        WHERE sku_id = %s
                        """
                        cursor.execute(query, (row['burnaby_qty'], sku_id))
                    elif update_kentucky:
                        query = """
                        UPDATE inventory_current
                        SET kentucky_qty = %s, last_updated = NOW()
                        WHERE sku_id = %s
                        """
                        cursor.execute(query, (row['kentucky_qty'], sku_id))

                    detail_entry["status"] = "success"
                    detail_entry["error_message"] = "Updated existing inventory record"
                    imported_count += 1
                    result["import_summary"]["successful"] += 1

                else:
                    # Insert new inventory record (only if both quantities provided)
                    if update_both:
                        query = """
                        INSERT INTO inventory_current (sku_id, burnaby_qty, kentucky_qty, last_updated)
                        VALUES (%s, %s, %s, NOW())
                        """
                        cursor.execute(query, (sku_id, row['burnaby_qty'], row['kentucky_qty']))

                        detail_entry["status"] = "success"
                        detail_entry["error_message"] = "Created new inventory record"
                        imported_count += 1
                        result["import_summary"]["successful"] += 1

                    else:
                        # Skip new SKU with partial data
                        detail_entry["status"] = "skipped"
                        detail_entry["error_category"] = "PARTIAL_DATA_NEW_SKU"
                        detail_entry["error_message"] = f"New SKU requires both warehouse quantities (has Burnaby: {update_burnaby}, has Kentucky: {update_kentucky})"
                        result["import_summary"]["skipped"] += 1

            except Exception as e:
                # Catch any database errors for individual SKUs
                detail_entry["status"] = "failed"
                detail_entry["error_category"] = "DATABASE_ERROR"
                detail_entry["error_message"] = f"Database error: {str(e)}"
                result["import_summary"]["failed"] += 1
                logger.error(f"Database error for SKU {sku_id} at line {line_number}: {str(e)}")

            result["import_details"].append(detail_entry)

        try:
            db.commit()
            db.close()

            result["records"] = imported_count

            # Log comprehensive summary
            summary = result["import_summary"]
            logger.info(f"Import completed for {sheet_name}: {summary['successful']} successful, {summary['failed']} failed, {summary['skipped']} skipped out of {summary['total_rows']} total rows")

            # Add warnings for any failures
            if summary["failed"] > 0:
                self.validation_warnings.append(f"{summary['failed']} rows failed to import - see detailed report")
            if summary["skipped"] > 0:
                self.validation_warnings.append(f"{summary['skipped']} rows skipped due to partial data")

        except Exception as e:
            self.validation_errors.append(f"Database commit error in {sheet_name}: {str(e)}")
            logger.error(f"Database commit error: {str(e)}")

        return result
    
    def _import_sales_data(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Import sales history data with stockout days"""
        
        expected_columns = ['sku_id', 'year_month', 'burnaby_sales', 'kentucky_sales', 'kentucky_stockout_days']
        result = {"type": "sales", "records": 0, "errors": []}
        
        # Validate columns (allow some flexibility)
        available_cols = df.columns.tolist()
        missing_critical = []
        
        for col in ['sku_id', 'year_month']:
            if col not in available_cols:
                missing_critical.append(col)
        
        if missing_critical:
            error_msg = f"Missing critical columns in {sheet_name}: {', '.join(missing_critical)}"
            self.validation_errors.append(error_msg)
            return result
        
        # Handle optional columns
        df_clean = df.copy()
        for col in ['burnaby_sales', 'kentucky_sales', 'kentucky_stockout_days', 'burnaby_stockout_days']:
            if col not in df_clean.columns:
                df_clean[col] = 0
        
        # Clean and validate data
        df_clean = df_clean.dropna(subset=['sku_id', 'year_month'])
        
        # Convert numeric columns
        numeric_cols = ['burnaby_sales', 'kentucky_sales', 'kentucky_stockout_days', 'burnaby_stockout_days']
        for col in numeric_cols:
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0).astype(int)
        
        # Validate stockout days (0-31)
        invalid_stockout = df_clean[
            (df_clean['kentucky_stockout_days'] < 0) | 
            (df_clean['kentucky_stockout_days'] > 31)
        ]
        
        if not invalid_stockout.empty:
            self.validation_warnings.append(
                f"Invalid stockout days found in {sheet_name} - should be 0-31 days"
            )
        
        # Import to database
        try:
            db = database.get_database_connection()
            cursor = db.cursor()
            
            imported_count = 0
            for _, row in df_clean.iterrows():
                # Upsert sales data
                query = """
                INSERT INTO monthly_sales
                (`sku_id`, `year_month`, `burnaby_sales`, `kentucky_sales`,
                 `burnaby_stockout_days`, `kentucky_stockout_days`)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    `burnaby_sales` = VALUES(`burnaby_sales`),
                    `kentucky_sales` = VALUES(`kentucky_sales`),
                    `burnaby_stockout_days` = VALUES(`burnaby_stockout_days`),
                    `kentucky_stockout_days` = VALUES(`kentucky_stockout_days`)
                """
                cursor.execute(query, (
                    row['sku_id'], row['year_month'], row['burnaby_sales'], 
                    row['kentucky_sales'], row['burnaby_stockout_days'], 
                    row['kentucky_stockout_days']
                ))
                imported_count += 1
            
            db.commit()
            db.close()
            
            result["records"] = imported_count
            logger.info(f"Imported {imported_count} sales records from {sheet_name}")
            
        except Exception as e:
            self.validation_errors.append(f"Database error in {sheet_name}: {str(e)}")
            logger.error(f"Database error importing sales: {str(e)}")
        
        return result
    
    def _import_sku_data(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """
        Import SKU master data with flexible column detection and batch processing

        Supports required and optional columns with intelligent detection.
        Uses batch processing for reliable import of large datasets.

        Args:
            df: DataFrame containing SKU data
            sheet_name: Name of the Excel sheet for error reporting

        Returns:
            Dict containing import results, record count, and any errors
        """

        # Required columns for SKU import
        required_columns = ['sku_id', 'description', 'supplier', 'cost_per_unit']

        # Optional columns that can be detected and imported
        optional_columns = ['status', 'transfer_multiple', 'abc_code', 'xyz_code', 'category']

        result = {"type": "sku", "records": 0, "errors": [], "warnings": []}

        # Validate required columns
        missing_cols = self._validate_required_columns(df, required_columns, sheet_name)
        if missing_cols:
            return result

        # Start with required columns
        columns_to_use = required_columns.copy()

        # Detect available optional columns using case-insensitive matching
        available_cols_lower = {col.lower(): col for col in df.columns}

        # Check for each optional column
        for opt_col in optional_columns:
            # Check for exact match first
            if opt_col in df.columns:
                columns_to_use.append(opt_col)
            # Check for case-insensitive match
            elif opt_col.lower() in available_cols_lower:
                columns_to_use.append(available_cols_lower[opt_col.lower()])

        # Create clean dataframe with detected columns
        df_clean = df[columns_to_use].copy()
        df_clean = df_clean.dropna(subset=['sku_id'])

        # Standardize column names (in case of case differences)
        column_mapping = {}
        for col in df_clean.columns:
            col_lower = col.lower()
            for opt_col in optional_columns:
                if opt_col.lower() == col_lower and col != opt_col:
                    column_mapping[col] = opt_col

        if column_mapping:
            df_clean = df_clean.rename(columns=column_mapping)

        # Add default values for missing optional columns
        if 'status' not in df_clean.columns:
            df_clean['status'] = 'Active'
        if 'transfer_multiple' not in df_clean.columns:
            df_clean['transfer_multiple'] = 50
        if 'abc_code' not in df_clean.columns:
            df_clean['abc_code'] = 'C'
        if 'xyz_code' not in df_clean.columns:
            df_clean['xyz_code'] = 'Z'
        if 'category' not in df_clean.columns:
            df_clean['category'] = None

        # Log which optional columns were detected
        detected_optional = [col for col in optional_columns if col in df_clean.columns and df_clean[col].notna().any()]
        if detected_optional:
            self.validation_warnings.append(f"Detected optional columns in {sheet_name}: {', '.join(detected_optional)}")
            logger.info(f"SKU import detected optional columns: {detected_optional}")

        # Import to database with batch processing
        total_rows = len(df_clean)
        batch_size = 500  # Process in batches of 500
        imported_count = 0
        failed_count = 0
        batch_errors = []

        try:
            db = database.get_database_connection()
            cursor = db.cursor()

            # Process in batches
            for batch_start in range(0, total_rows, batch_size):
                batch_end = min(batch_start + batch_size, total_rows)
                batch_df = df_clean.iloc[batch_start:batch_end]

                try:
                    # Start transaction for this batch
                    db.begin()

                    # Prepare batch data
                    batch_data = []
                    for _, row in batch_df.iterrows():
                        batch_data.append((
                            row['sku_id'],
                            row['description'],
                            row['supplier'],
                            row['cost_per_unit'],
                            row.get('status', 'Active'),
                            row.get('transfer_multiple', 50),
                            row.get('abc_code', 'C'),
                            row.get('xyz_code', 'Z'),
                            row.get('category', None)
                        ))

                    # Execute batch insert
                    query = """
                    INSERT INTO skus (sku_id, description, supplier, cost_per_unit, status, transfer_multiple, abc_code, xyz_code, category)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        description = VALUES(description),
                        supplier = VALUES(supplier),
                        cost_per_unit = VALUES(cost_per_unit),
                        status = VALUES(status),
                        transfer_multiple = VALUES(transfer_multiple),
                        abc_code = VALUES(abc_code),
                        xyz_code = VALUES(xyz_code),
                        category = VALUES(category)
                    """

                    cursor.executemany(query, batch_data)
                    db.commit()

                    batch_imported = len(batch_data)
                    imported_count += batch_imported

                    logger.info(f"Batch {batch_start//batch_size + 1}: Successfully imported {batch_imported} SKUs (rows {batch_start+1}-{batch_end})")

                except Exception as batch_error:
                    # Rollback failed batch
                    db.rollback()
                    failed_count += len(batch_df)
                    error_msg = f"Batch {batch_start//batch_size + 1} failed (rows {batch_start+1}-{batch_end}): {str(batch_error)}"
                    batch_errors.append(error_msg)
                    logger.error(error_msg)

                    # Try individual inserts for this batch to identify specific issues
                    individual_failures = []
                    for _, row in batch_df.iterrows():
                        try:
                            cursor.execute(query, (
                                row['sku_id'],
                                row['description'],
                                row['supplier'],
                                row['cost_per_unit'],
                                row.get('status', 'Active'),
                                row.get('transfer_multiple', 50),
                                row.get('abc_code', 'C'),
                                row.get('xyz_code', 'Z'),
                                row.get('category', None)
                            ))
                            db.commit()
                            imported_count += 1
                        except Exception as individual_error:
                            individual_failures.append(f"SKU {row['sku_id']}: {str(individual_error)}")

                    if individual_failures:
                        batch_errors.extend(individual_failures[:5])  # Limit to first 5 errors
                        if len(individual_failures) > 5:
                            batch_errors.append(f"... and {len(individual_failures) - 5} more errors in this batch")

            cursor.close()
            db.close()

            # Compile results
            result["records"] = imported_count
            result["errors"] = batch_errors

            if imported_count > 0:
                logger.info(f"Successfully imported {imported_count} SKU records from {sheet_name}")
                if failed_count > 0:
                    logger.warning(f"Failed to import {failed_count} SKU records from {sheet_name}")
            else:
                logger.error(f"Failed to import any SKU records from {sheet_name}")

            # Add summary to warnings
            if batch_errors:
                self.validation_warnings.append(f"Import completed with errors: {imported_count} successful, {failed_count} failed")

        except Exception as e:
            error_msg = f"Critical database error in {sheet_name}: {str(e)}"
            logger.error(error_msg)
            self.validation_errors.append(error_msg)
            if 'db' in locals():
                try:
                    db.rollback()
                    db.close()
                except:
                    pass

        return result
    
    def _auto_detect_and_import(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Auto-detect sheet content type and import appropriately"""

        columns = [col.lower() for col in df.columns]

        # Check for inventory data - requires sku_id and at least one quantity column
        has_sku_id = 'sku_id' in columns
        has_burnaby_qty = any('burnaby' in col and 'qty' in col for col in columns)
        has_kentucky_qty = any('kentucky' in col and 'qty' in col for col in columns)

        if has_sku_id and (has_burnaby_qty or has_kentucky_qty):
            return self._import_inventory_data(df, sheet_name)
        elif 'year_month' in columns and any('sales' in col for col in columns):
            return self._import_sales_data(df, sheet_name)
        elif 'description' in columns and 'cost_per_unit' in columns:
            return self._import_sku_data(df, sheet_name)
        else:
            self.validation_warnings.append(f"Unknown sheet format: {sheet_name} - skipped")
            return {"type": "unknown", "records": 0}
    
    def _validate_required_columns(self, df: pd.DataFrame, required_cols: List[str], sheet_name: str) -> List[str]:
        """Validate that required columns exist in dataframe"""
        
        available_cols = [col.lower() for col in df.columns]
        missing_cols = []
        
        for col in required_cols:
            if col.lower() not in available_cols:
                missing_cols.append(col)
        
        if missing_cols:
            error_msg = f"Missing required columns in {sheet_name}: {', '.join(missing_cols)}"
            self.validation_errors.append(error_msg)
        
        return missing_cols
    
    def _generate_import_summary(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive import summary"""
        
        return {
            "total_sheets": results["sheets_processed"],
            "total_records": results["records_imported"],
            "error_count": len(results["errors"]),
            "warning_count": len(results["warnings"]),
            "success_rate": "100%" if results["success"] else f"{max(0, 100 - len(results['errors']) * 10)}%"
        }
    
    # =============================================================================
    # EXCEL EXPORT FUNCTIONALITY
    # =============================================================================
    
    def export_transfer_recommendations_excel(self, recommendations: List[Dict]) -> bytes:
        """
        Export transfer recommendations to professionally formatted Excel file with pending orders

        Args:
            recommendations: List of transfer recommendation dictionaries

        Returns:
            bytes: Excel file content ready for download
        """

        # Create workbook with multiple sheets
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Transfer Orders sheet with enhanced data
        self._create_enhanced_transfer_orders_sheet(wb, recommendations)

        # Create Pending Orders sheet
        self._create_pending_orders_sheet(wb)

        # Create Coverage Analysis sheet
        self._create_coverage_analysis_sheet(wb)

        # Create Summary sheet
        self._create_summary_sheet(wb, recommendations)

        # Create Inventory Status sheet
        self._create_inventory_status_sheet(wb)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
    
    def _create_enhanced_transfer_orders_sheet(self, wb: Workbook, recommendations: List[Dict]):
        """Create enhanced transfer orders sheet with pending orders and coverage analysis"""

        ws = wb.create_sheet("Transfer Orders")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        priority_fills = {
            "CRITICAL": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
            "HIGH": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
            "LOW": PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        }

        # Enhanced headers with pending orders and coverage data
        headers = [
            "SKU", "Description", "Priority", "Current KY Qty", "Available CA Qty",
            "Pending CA", "Pending KY", "CA Coverage After", "KY Coverage After",
            "Monthly Demand", "Coverage (Months)", "Recommended Transfer",
            "ABC/XYZ Class", "Stockout Override", "Reason", "Transfer Multiple"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Get pending orders data
        pending_data = self._get_pending_orders_summary()

        # Write data with enhanced information
        for row, rec in enumerate(recommendations, 2):
            sku_id = rec['sku_id']

            ws.cell(row=row, column=1, value=sku_id)
            ws.cell(row=row, column=2, value=rec['description'])

            # Priority with color coding
            priority_cell = ws.cell(row=row, column=3, value=rec['priority'])
            if rec['priority'] in priority_fills:
                priority_cell.fill = priority_fills[rec['priority']]
                if rec['priority'] in ['CRITICAL', 'HIGH', 'LOW']:
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                else:
                    priority_cell.font = Font(bold=True)

            ws.cell(row=row, column=4, value=rec['current_kentucky_qty'])
            ws.cell(row=row, column=5, value=rec['current_burnaby_qty'])

            # Pending orders data
            pending_ca = pending_data.get(sku_id, {}).get('burnaby_pending', 0)
            pending_ky = pending_data.get(sku_id, {}).get('kentucky_pending', 0)
            ws.cell(row=row, column=6, value=pending_ca)
            ws.cell(row=row, column=7, value=pending_ky)

            # Coverage after transfer calculations
            monthly_demand = rec.get('corrected_monthly_demand', 0)
            transfer_qty = rec.get('recommended_transfer_qty', 0)

            if monthly_demand > 0:
                ca_coverage_after = (rec['current_burnaby_qty'] - transfer_qty + pending_ca) / monthly_demand
                ky_coverage_after = (rec['current_kentucky_qty'] + transfer_qty + pending_ky) / monthly_demand
            else:
                ca_coverage_after = 0
                ky_coverage_after = 0

            ws.cell(row=row, column=8, value=round(ca_coverage_after, 1))
            ws.cell(row=row, column=9, value=round(ky_coverage_after, 1))

            ws.cell(row=row, column=10, value=round(monthly_demand))
            ws.cell(row=row, column=11, value=round(rec['coverage_months'], 1))
            ws.cell(row=row, column=12, value=transfer_qty)
            ws.cell(row=row, column=13, value=f"{rec['abc_class']}{rec['xyz_class']}")

            # Stockout override indicator
            stockout_override = "YES" if rec.get('stockout_override_applied', False) else "NO"
            override_cell = ws.cell(row=row, column=14, value=stockout_override)
            if stockout_override == "YES":
                override_cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                override_cell.font = Font(bold=True)

            ws.cell(row=row, column=15, value=rec['reason'])
            ws.cell(row=row, column=16, value=rec['transfer_multiple'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(recommendations)+1):
            for cell in row:
                cell.border = thin_border

    def _create_transfer_orders_sheet(self, wb: Workbook, recommendations: List[Dict]):
        """Create the main transfer orders sheet with professional formatting (legacy method)"""
        
        ws = wb.create_sheet("Transfer Orders")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        priority_fills = {
            "CRITICAL": PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
            "HIGH": PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
            "LOW": PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        }
        
        # Headers
        headers = [
            "SKU", "Description", "Priority", "Current KY Qty", "Available CA Qty",
            "Monthly Demand", "Coverage (Months)", "Recommended Transfer", 
            "ABC/XYZ Class", "Reason", "Transfer Multiple"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row, rec in enumerate(recommendations, 2):
            ws.cell(row=row, column=1, value=rec['sku_id'])
            ws.cell(row=row, column=2, value=rec['description'])
            
            # Priority with color coding
            priority_cell = ws.cell(row=row, column=3, value=rec['priority'])
            if rec['priority'] in priority_fills:
                priority_cell.fill = priority_fills[rec['priority']]
                if rec['priority'] in ['CRITICAL', 'HIGH', 'LOW']:
                    priority_cell.font = Font(color="FFFFFF", bold=True)
                else:
                    priority_cell.font = Font(bold=True)
            
            ws.cell(row=row, column=4, value=rec['current_kentucky_qty'])
            ws.cell(row=row, column=5, value=rec['current_burnaby_qty'])
            ws.cell(row=row, column=6, value=round(rec['corrected_monthly_demand']))
            ws.cell(row=row, column=7, value=round(rec['coverage_months'], 1))
            ws.cell(row=row, column=8, value=rec['recommended_transfer_qty'])
            ws.cell(row=row, column=9, value=f"{rec['abc_class']}{rec['xyz_class']}")
            ws.cell(row=row, column=10, value=rec['reason'])
            ws.cell(row=row, column=11, value=rec['transfer_multiple'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(recommendations)+1):
            for cell in row:
                cell.border = thin_border
    
    def _create_summary_sheet(self, wb: Workbook, recommendations: List[Dict]):
        """Create summary statistics sheet"""
        
        ws = wb.create_sheet("Summary")
        
        # Calculate statistics
        total_items = len(recommendations)
        critical_items = len([r for r in recommendations if r['priority'] == 'CRITICAL'])
        high_items = len([r for r in recommendations if r['priority'] == 'HIGH'])
        total_transfer_qty = sum(r['recommended_transfer_qty'] for r in recommendations)
        avg_coverage = sum(r['coverage_months'] for r in recommendations) / total_items if total_items > 0 else 0
        
        # Summary data
        summary_data = [
            ["Transfer Planning Summary", ""],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Total SKUs Analyzed", total_items],
            ["Critical Priority Items", critical_items],
            ["High Priority Items", high_items],
            ["Total Transfer Quantity", total_transfer_qty],
            ["Average Coverage (Months)", round(avg_coverage, 1)],
            ["", ""],
            ["Priority Distribution", "Count"],
        ]
        
        # Count by priority
        priority_counts = {}
        for rec in recommendations:
            priority = rec['priority']
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
        
        for priority, count in priority_counts.items():
            summary_data.append([f"  {priority}", count])
        
        # Write summary data
        for row, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            
            # Format header row
            if row == 1:
                ws.cell(row=row, column=1).font = Font(bold=True, size=14)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=14)
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_inventory_status_sheet(self, wb: Workbook):
        """Create current inventory status sheet"""
        
        ws = wb.create_sheet("Current Inventory")
        
        try:
            # Get current inventory data
            db = database.get_database_connection()
            cursor = db.cursor(pymysql.cursors.DictCursor)
            
            query = """
            SELECT 
                s.sku_id,
                s.description,
                s.supplier,
                ic.burnaby_qty,
                ic.kentucky_qty,
                (ic.burnaby_qty + ic.kentucky_qty) as total_qty,
                s.cost_per_unit,
                (ic.kentucky_qty * s.cost_per_unit) as kentucky_value,
                ic.last_updated
            FROM skus s
            LEFT JOIN inventory_current ic ON s.sku_id = ic.sku_id
            WHERE s.status = 'Active'
            ORDER BY ic.kentucky_qty ASC, s.sku_id
            """
            
            cursor.execute(query)
            inventory_data = cursor.fetchall()
            db.close()
            
            # Headers
            headers = ["SKU", "Description", "Supplier", "Burnaby Qty", "Kentucky Qty", 
                      "Total Qty", "Unit Cost", "KY Value", "Last Updated"]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row, item in enumerate(inventory_data, 2):
                for col, header in enumerate(headers, 1):
                    key = header.lower().replace(' ', '_').replace('#', '').replace('/', '_')
                    if key == 'ky_value':
                        key = 'kentucky_value'
                    
                    value = item.get(key, '')
                    ws.cell(row=row, column=col, value=value)
                    
                    # Color code out of stock items
                    if col == 5 and value == 0:  # Kentucky Qty column
                        for c in range(1, len(headers) + 1):
                            ws.cell(row=row, column=c).fill = PatternFill(
                                start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                            )
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            ws.cell(row=1, column=1, value=f"Error loading inventory data: {str(e)}")
            logger.error(f"Error creating inventory status sheet: {str(e)}")
    
    # =============================================================================
    # CSV IMPORT/EXPORT FUNCTIONALITY
    # =============================================================================
    
    def export_csv(self, data: List[Dict], filename_prefix: str = "export") -> bytes:
        """Export data to CSV format"""
        
        if not data:
            return b"No data to export"
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Convert to CSV
        output = io.StringIO()
        df.to_csv(output, index=False)
        
        return output.getvalue().encode('utf-8')
    
    def import_csv_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Import CSV file with comprehensive validation and detailed reporting"""

        try:
            # Reset validation tracking
            self.validation_errors = []
            self.validation_warnings = []

            # Read CSV
            df = pd.read_csv(io.StringIO(file_content.decode('utf-8')))

            # Create structured results similar to Excel import
            results = {
                "success": False,
                "filename": filename,
                "import_timestamp": datetime.now().isoformat(),
                "sheets_processed": 1,
                "records_imported": 0,
                "errors": [],
                "warnings": [],
                "results": []  # Add results array for detailed import data
            }

            # Process the CSV (treat as single sheet)
            sheet_result = self._auto_detect_and_import(df, filename)
            results["records_imported"] = sheet_result.get("records", 0)
            results["results"].append(sheet_result)  # Add detailed results

            # Compile validation results
            results["errors"] = self.validation_errors
            results["warnings"] = self.validation_warnings
            results["success"] = len(self.validation_errors) == 0

            logger.info(f"CSV import completed: {filename}, Success: {results['success']}, Records: {results['records_imported']}")
            return results

        except Exception as e:
            logger.error(f"CSV import failed: {str(e)}")
            return {
                "success": False,
                "filename": filename,
                "error": f"CSV import failed: {str(e)}",
                "import_timestamp": datetime.now().isoformat()
            }

    def import_pending_orders_csv(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Import pending orders from CSV with flexible date handling

        Supports flexible CSV formats:
        - Required: sku_id, quantity, destination
        - Optional: expected_date, order_date, order_type, notes
        - Auto-calculates dates when missing (Today + 120 days for expected_arrival)

        Args:
            file_content: Raw CSV file bytes
            filename: Original filename for error reporting

        Returns:
            Dict with import results, validation errors, and statistics
        """
        try:
            # Reset validation tracking
            self.validation_errors = []
            self.validation_warnings = []

            # Read CSV
            df = pd.read_csv(io.StringIO(file_content.decode('utf-8')))

            # Required columns
            required_columns = ['sku_id', 'quantity', 'destination']
            optional_columns = ['expected_date', 'expected_arrival', 'order_date', 'order_type', 'notes', 'lead_time_days']

            result = {
                "success": False,
                "filename": filename,
                "import_timestamp": datetime.now().isoformat(),
                "records_imported": 0,
                "errors": [],
                "warnings": [],
                "summary": {}
            }

            # Validate required columns (case-insensitive)
            df_cols_lower = {col.lower(): col for col in df.columns}
            missing_cols = []

            for req_col in required_columns:
                if req_col.lower() not in df_cols_lower:
                    missing_cols.append(req_col)

            if missing_cols:
                error_msg = f"Missing required columns: {', '.join(missing_cols)}"
                self.validation_errors.append(error_msg)
                result["errors"] = self.validation_errors
                return result

            # Map columns to standard names (case-insensitive)
            column_mapping = {}
            for col in df.columns:
                col_lower = col.lower()
                if col_lower in [req.lower() for req in required_columns + optional_columns]:
                    # Find the standard name
                    for std_col in required_columns + optional_columns:
                        if col_lower == std_col.lower():
                            if col != std_col:
                                column_mapping[col] = std_col
                            break

            # Rename columns
            if column_mapping:
                df = df.rename(columns=column_mapping)

            # Clean data
            df_clean = df.dropna(subset=['sku_id'])
            df_clean = df_clean[df_clean['sku_id'].str.strip() != '']

            # Validate destinations
            valid_destinations = ['burnaby', 'kentucky']
            df_clean['destination'] = df_clean['destination'].str.lower().str.strip()
            invalid_destinations = df_clean[~df_clean['destination'].isin(valid_destinations)]

            if not invalid_destinations.empty:
                self.validation_warnings.append(
                    f"Invalid destinations found (will be skipped): {', '.join(invalid_destinations['destination'].unique())}"
                )
                df_clean = df_clean[df_clean['destination'].isin(valid_destinations)]

            # Validate quantities
            df_clean['quantity'] = pd.to_numeric(df_clean['quantity'], errors='coerce')
            df_clean = df_clean.dropna(subset=['quantity'])
            df_clean = df_clean[df_clean['quantity'] > 0]

            # Handle dates with flexibility
            today = datetime.now().date()
            default_lead_time = 120  # days

            # Set defaults for missing optional columns
            if 'order_date' not in df_clean.columns:
                df_clean['order_date'] = today

            if 'lead_time_days' not in df_clean.columns:
                df_clean['lead_time_days'] = default_lead_time

            if 'order_type' not in df_clean.columns:
                df_clean['order_type'] = 'supplier'

            if 'notes' not in df_clean.columns:
                df_clean['notes'] = None

            # Handle expected arrival date with enhanced validation for mixed imports
            expected_date_col = None
            potential_date_cols = ['expected_date', 'expected_arrival', 'arrival_date', 'expected arrival', 'arrival date']

            for col in potential_date_cols:
                # Check both exact and case-insensitive matches
                if col in df_clean.columns:
                    expected_date_col = col
                    break
                elif col.lower() in [c.lower() for c in df_clean.columns]:
                    # Find the actual column name (case-insensitive)
                    for actual_col in df_clean.columns:
                        if actual_col.lower() == col.lower():
                            expected_date_col = actual_col
                            break
                    if expected_date_col:
                        break

            # Enhanced date handling with detailed validation feedback
            if expected_date_col:
                self.validation_warnings.append(f"Found expected date column: '{expected_date_col}'")

                # Parse existing expected dates with enhanced validation
                original_values = df_clean[expected_date_col].copy()
                df_clean['expected_arrival'] = pd.to_datetime(df_clean[expected_date_col], errors='coerce').dt.date

                # Count and report different date scenarios
                valid_dates = ~df_clean['expected_arrival'].isna()
                invalid_dates = df_clean[expected_date_col].notna() & df_clean['expected_arrival'].isna()
                missing_dates = df_clean[expected_date_col].isna()

                valid_count = valid_dates.sum()
                invalid_count = invalid_dates.sum()
                missing_count = missing_dates.sum()

                # Provide detailed feedback
                if valid_count > 0:
                    self.validation_warnings.append(f"Successfully parsed {valid_count} expected dates")

                if invalid_count > 0:
                    invalid_examples = original_values[invalid_dates].dropna().unique()[:3]
                    self.validation_warnings.append(
                        f"Found {invalid_count} invalid date formats (will use default): {', '.join(map(str, invalid_examples))}"
                    )

                if missing_count > 0:
                    self.validation_warnings.append(f"Found {missing_count} empty date fields (will use default)")

                # Fill missing or invalid dates with calculated date
                needs_default = df_clean['expected_arrival'].isna()
                # Use order_date + lead_time instead of today's date for estimated arrivals
                df_clean.loc[needs_default, 'expected_arrival'] = pd.to_datetime(df_clean.loc[needs_default, 'order_date']).dt.date + pd.Timedelta(days=default_lead_time)
                df_clean.loc[needs_default, 'is_estimated'] = True

                # Mark provided dates as confirmed
                df_clean.loc[~needs_default, 'is_estimated'] = False

            else:
                # No expected date column - calculate all using order_date + lead_time
                df_clean['expected_arrival'] = pd.to_datetime(df_clean['order_date']).dt.date + pd.Timedelta(days=default_lead_time)
                df_clean['is_estimated'] = True
                self.validation_warnings.append("No expected date column found - using order_date + 120 days for all orders")

            # Convert quantities to integers
            df_clean['quantity'] = df_clean['quantity'].astype(int)
            df_clean['lead_time_days'] = pd.to_numeric(df_clean['lead_time_days'], errors='coerce').fillna(default_lead_time).astype(int)

            # Validate SKUs exist in database
            try:
                db = database.get_database_connection()
                cursor = db.cursor()

                # Get list of valid SKUs
                cursor.execute("SELECT sku_id FROM skus WHERE status = 'Active'")
                valid_skus = {row[0] for row in cursor.fetchall()}

                # Filter out invalid SKUs
                invalid_skus = df_clean[~df_clean['sku_id'].isin(valid_skus)]['sku_id'].unique()
                if len(invalid_skus) > 0:
                    self.validation_warnings.append(
                        f"Invalid SKUs found (will be skipped): {', '.join(invalid_skus[:5])}" +
                        (f" and {len(invalid_skus)-5} more" if len(invalid_skus) > 5 else "")
                    )
                    df_clean = df_clean[df_clean['sku_id'].isin(valid_skus)]

                # Insert records
                imported_count = 0
                for _, row in df_clean.iterrows():
                    insert_query = """
                    INSERT INTO pending_inventory
                    (sku_id, quantity, destination, expected_arrival, order_date, lead_time_days, order_type, notes, is_estimated, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
                    """
                    cursor.execute(insert_query, (
                        row['sku_id'],
                        row['quantity'],
                        row['destination'],
                        row['expected_arrival'],
                        row.get('order_date', today),
                        row.get('lead_time_days', default_lead_time),
                        row.get('order_type', 'supplier'),
                        row.get('notes'),
                        bool(row.get('is_estimated', True))  # Use calculated is_estimated flag
                    ))
                    imported_count += 1

                db.commit()
                db.close()

                # Success
                result["success"] = True
                result["records_imported"] = imported_count
                result["errors"] = self.validation_errors
                result["warnings"] = self.validation_warnings
                # Enhanced summary with detailed import statistics
                estimated_dates_count = len(df_clean[df_clean['is_estimated'] == True])
                confirmed_dates_count = len(df_clean[df_clean['is_estimated'] == False])

                result["summary"] = {
                    "total_records": imported_count,
                    "skus_processed": len(df_clean['sku_id'].unique()),
                    "destinations": list(df_clean['destination'].unique()),
                    "confirmed_dates": confirmed_dates_count,
                    "estimated_dates": estimated_dates_count,
                    "date_column_found": expected_date_col is not None,
                    "date_column_name": expected_date_col if expected_date_col else None,
                    "mixed_import": expected_date_col is not None and estimated_dates_count > 0 and confirmed_dates_count > 0,
                    "lead_time_used": default_lead_time
                }

                logger.info(f"Successfully imported {imported_count} pending orders from {filename}")
                return result

            except Exception as e:
                db.rollback() if 'db' in locals() else None
                db.close() if 'db' in locals() else None
                self.validation_errors.append(f"Database error: {str(e)}")
                logger.error(f"Database error importing pending orders: {str(e)}")
                result["errors"] = self.validation_errors
                return result

        except Exception as e:
            logger.error(f"Pending orders CSV import failed: {str(e)}")
            return {
                "success": False,
                "filename": filename,
                "error": f"Import failed: {str(e)}",
                "import_timestamp": datetime.now().isoformat()
            }

    def _get_pending_orders_summary(self) -> Dict[str, Dict]:
        """
        Get summary of pending orders grouped by SKU and destination

        Returns:
            Dict[sku_id, Dict]: Pending orders data by SKU
        """
        try:
            db = database.get_database_connection()
            cursor = db.cursor(pymysql.cursors.DictCursor)

            query = """
            SELECT
                sku_id,
                destination,
                SUM(quantity) as total_quantity,
                COUNT(*) as order_count,
                MIN(expected_arrival) as earliest_arrival,
                MAX(expected_arrival) as latest_arrival,
                AVG(DATEDIFF(expected_arrival, CURDATE())) as avg_days_until_arrival
            FROM pending_inventory
            WHERE status IN ('ordered', 'shipped', 'pending')
            GROUP BY sku_id, destination
            """

            cursor.execute(query)
            results = cursor.fetchall()

            # Organize data by SKU
            summary = {}
            for row in results:
                sku_id = row['sku_id']
                if sku_id not in summary:
                    summary[sku_id] = {'burnaby_pending': 0, 'kentucky_pending': 0}

                if row['destination'] == 'burnaby':
                    summary[sku_id]['burnaby_pending'] = row['total_quantity']
                    summary[sku_id]['burnaby_earliest'] = row['earliest_arrival']
                elif row['destination'] == 'kentucky':
                    summary[sku_id]['kentucky_pending'] = row['total_quantity']
                    summary[sku_id]['kentucky_earliest'] = row['earliest_arrival']

            cursor.close()
            db.close()
            return summary

        except Exception as e:
            logger.error(f"Error getting pending orders summary: {str(e)}")
            return {}

    def _create_pending_orders_sheet(self, wb: Workbook):
        """Create pending orders sheet showing all current pending orders"""

        ws = wb.create_sheet("Pending Orders")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Headers for pending orders
        headers = [
            "SKU", "Description", "Supplier", "Quantity", "Destination",
            "Order Date", "Expected Arrival", "Days Until Arrival",
            "Lead Time (Days)", "Order Type", "Status", "Is Estimated", "Notes"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        try:
            db = database.get_database_connection()
            cursor = db.cursor(pymysql.cursors.DictCursor)

            query = """
            SELECT
                pi.*,
                s.description,
                s.supplier,
                DATEDIFF(pi.expected_arrival, CURDATE()) as days_until_arrival
            FROM pending_inventory pi
            LEFT JOIN skus s ON pi.sku_id = s.sku_id
            WHERE pi.status IN ('ordered', 'shipped', 'pending')
            ORDER BY pi.expected_arrival ASC, pi.sku_id
            """

            cursor.execute(query)
            pending_orders = cursor.fetchall()

            # Status color coding
            status_fills = {
                "ordered": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                "shipped": PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"),
                "pending": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            }

            # Write data
            for row_idx, order in enumerate(pending_orders, 2):
                ws.cell(row=row_idx, column=1, value=order['sku_id'])
                ws.cell(row=row_idx, column=2, value=order['description'])
                ws.cell(row=row_idx, column=3, value=order['supplier'])
                ws.cell(row=row_idx, column=4, value=order['quantity'])
                ws.cell(row=row_idx, column=5, value=order['destination'].title())
                ws.cell(row=row_idx, column=6, value=order['order_date'])
                ws.cell(row=row_idx, column=7, value=order['expected_arrival'])
                ws.cell(row=row_idx, column=8, value=order['days_until_arrival'])
                ws.cell(row=row_idx, column=9, value=order['lead_time_days'])
                ws.cell(row=row_idx, column=10, value=order['order_type'].title())

                # Status cell with color coding
                status_cell = ws.cell(row=row_idx, column=11, value=order['status'].title())
                if order['status'] in status_fills:
                    status_cell.fill = status_fills[order['status']]

                estimated_text = "Yes" if order['is_estimated'] else "No"
                estimated_cell = ws.cell(row=row_idx, column=12, value=estimated_text)
                if order['is_estimated']:
                    estimated_cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

                ws.cell(row=row_idx, column=13, value=order['notes'] or "")

            cursor.close()
            db.close()

        except Exception as e:
            logger.error(f"Error creating pending orders sheet: {str(e)}")
            # Add error row
            ws.cell(row=2, column=1, value=f"Error loading data: {str(e)}")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_coverage_analysis_sheet(self, wb: Workbook):
        """Create coverage analysis sheet showing current vs projected coverage"""

        ws = wb.create_sheet("Coverage Analysis")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Headers
        headers = [
            "SKU", "Description", "ABC/XYZ", "Monthly Demand",
            "Current CA Qty", "Current KY Qty", "Pending CA", "Pending KY",
            "Current CA Coverage", "Current KY Coverage", "Projected CA Coverage", "Projected KY Coverage",
            "Coverage Improvement CA", "Coverage Improvement KY", "Status"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        try:
            db = database.get_database_connection()
            cursor = db.cursor(pymysql.cursors.DictCursor)

            # Get inventory and demand data with pending orders
            query = """
            SELECT
                s.sku_id,
                s.description,
                CONCAT(s.abc_code, s.xyz_code) as abc_xyz,
                COALESCE(ms.burnaby_sales, 0) + COALESCE(ms.kentucky_sales, 0) as monthly_demand,
                COALESCE(ic.burnaby_qty, 0) as current_burnaby_qty,
                COALESCE(ic.kentucky_qty, 0) as current_kentucky_qty,
                COALESCE(pb.pending_burnaby, 0) as pending_burnaby,
                COALESCE(pk.pending_kentucky, 0) as pending_kentucky
            FROM skus s
            LEFT JOIN inventory_current ic ON s.sku_id = ic.sku_id
            LEFT JOIN monthly_sales ms ON s.sku_id = ms.sku_id
            LEFT JOIN (
                SELECT sku_id, SUM(quantity) as pending_burnaby
                FROM pending_inventory
                WHERE destination = 'burnaby' AND status IN ('ordered', 'shipped', 'pending')
                GROUP BY sku_id
            ) pb ON s.sku_id = pb.sku_id
            LEFT JOIN (
                SELECT sku_id, SUM(quantity) as pending_kentucky
                FROM pending_inventory
                WHERE destination = 'kentucky' AND status IN ('ordered', 'shipped', 'pending')
                GROUP BY sku_id
            ) pk ON s.sku_id = pk.sku_id
            WHERE s.status = 'Active'
            ORDER BY s.sku_id
            """

            cursor.execute(query)
            results = cursor.fetchall()

            # Write data with calculations
            for row_idx, row in enumerate(results, 2):
                monthly_demand = max(row['monthly_demand'], 0.1)  # Avoid division by zero

                ws.cell(row=row_idx, column=1, value=row['sku_id'])
                ws.cell(row=row_idx, column=2, value=row['description'])
                ws.cell(row=row_idx, column=3, value=row['abc_xyz'])
                ws.cell(row=row_idx, column=4, value=round(monthly_demand))

                current_ca = row['current_burnaby_qty']
                current_ky = row['current_kentucky_qty']
                pending_ca = row['pending_burnaby']
                pending_ky = row['pending_kentucky']

                ws.cell(row=row_idx, column=5, value=current_ca)
                ws.cell(row=row_idx, column=6, value=current_ky)
                ws.cell(row=row_idx, column=7, value=pending_ca)
                ws.cell(row=row_idx, column=8, value=pending_ky)

                # Current coverage
                current_ca_coverage = current_ca / monthly_demand
                current_ky_coverage = current_ky / monthly_demand

                # Projected coverage (with pending orders)
                projected_ca_coverage = (current_ca + pending_ca) / monthly_demand
                projected_ky_coverage = (current_ky + pending_ky) / monthly_demand

                ws.cell(row=row_idx, column=9, value=round(current_ca_coverage, 1))
                ws.cell(row=row_idx, column=10, value=round(current_ky_coverage, 1))
                ws.cell(row=row_idx, column=11, value=round(projected_ca_coverage, 1))
                ws.cell(row=row_idx, column=12, value=round(projected_ky_coverage, 1))

                # Coverage improvements
                ca_improvement = projected_ca_coverage - current_ca_coverage
                ky_improvement = projected_ky_coverage - current_ky_coverage

                improvement_ca_cell = ws.cell(row=row_idx, column=13, value=round(ca_improvement, 1))
                improvement_ky_cell = ws.cell(row=row_idx, column=14, value=round(ky_improvement, 1))

                # Color code improvements
                if ca_improvement > 0:
                    improvement_ca_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                if ky_improvement > 0:
                    improvement_ky_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                # Overall status
                if pending_ca + pending_ky > 0:
                    if projected_ky_coverage < 1.0:  # Still low coverage
                        status = "LOW COVERAGE"
                        status_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    else:
                        status = "IMPROVING"
                        status_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                else:
                    status = "NO PENDING"
                    status_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                status_cell = ws.cell(row=row_idx, column=15, value=status)
                status_cell.fill = status_fill

            cursor.close()
            db.close()

        except Exception as e:
            logger.error(f"Error creating coverage analysis sheet: {str(e)}")
            ws.cell(row=2, column=1, value=f"Error loading data: {str(e)}")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"


# Global instance
import_export_manager = ImportExportManager()